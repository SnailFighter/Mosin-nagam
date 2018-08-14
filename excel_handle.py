from openpyxl import Workbook


def export_excel(data, product_name):

    wb = Workbook()
    # grab the active worksheet
    ws = wb.active

    # Data can be assigned directly to cells
    ws['A1'] = "关键词"
    ws['B1'] = "次数"
    ws['C1'] = "公司名称"
    ws['E1'] = "公司阿里网址"

    index = 2
    # 排版数据
    for (k, v) in data.items():
        print("-----排版数据start-------")
        ws["A"+str(index)] = v["key"]
        ws["B" + str(index)] = v["time"]
        ws["C" + str(index)] = v["company"]
        ws["E" + str(index)] = v["url"]
        index = index+1
        print("-----排版数据end-------")

    wb.save("阿里巴巴国际版关键词_"+product_name+".xlsx")
